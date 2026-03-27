# app.py
import os
from flask import Flask, request, redirect, url_for, flash, session, send_from_directory
from models import create_all, create_default_admin_if_not_exists, User, Customer, WorkOrder, OrderImage, OrderHistory
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
from sqlalchemy.orm import sessionmaker
from models import get_engine
from sqlalchemy import case, desc
from flask import render_template
from sqlalchemy.orm import joinedload
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file
from models import TechnicianSkill


def build_monthly_data(orders):
    data = defaultdict(int)

    for o in orders:
        if o.created_at:
            key = o.created_at.strftime("%m/%Y")
            data[key] += 1

    if not data:
        return json.dumps({
            "labels": ["Chưa có dữ liệu"],
            "datasets": [{
                "label": "Số đơn",
                "data": [0]
            }]
        })

    labels = sorted(data.keys())

    return json.dumps({
        "labels": labels,
        "datasets": [{
            "label": "Số đơn",
            "data": [data[l] for l in labels]
        }]
    })




def build_type_data(orders):
    bao_tri = sum(1 for o in orders if getattr(o, "order_type", None) == "bao_tri")
    sua_chua = sum(1 for o in orders if getattr(o, "order_type", None) == "sua_chua")

    if bao_tri == 0 and sua_chua == 0:
        return json.dumps({
            "labels": ["Chưa có dữ liệu"],
            "datasets": [{"data": [1]}]
        })

    return json.dumps({
        "labels": ["Bảo trì", "Sửa chữa"],
        "datasets": [{
            "data": [bao_tri, sua_chua]
        }]
    })


def build_tech_data(orders):
    from collections import defaultdict
    import json

    data = defaultdict(int)

    for o in orders:
        if getattr(o, "technician", None):
            data[o.technician.name] += 1

    if not data:
        return json.dumps({
            "labels": ["Chưa phân công"],
            "datasets": [{
                "label": "Số đơn",
                "data": [0]
            }]
        })

    return json.dumps({
        "labels": list(data.keys()),
        "datasets": [{
            "label": "Số đơn",
            "data": list(data.values())
        }]
    })




engine = get_engine()
SessionLocal = sessionmaker(bind=engine)


UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
ALLOWED_EXT = {'png','jpg','jpeg','gif'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__)
app.secret_key = 'change_me_now'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024

engine = create_all()
Session = sessionmaker(bind=engine)


s = Session()
create_default_admin_if_not_exists(s)
s.close()

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper

def role_required(*roles):
    def deco(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if session.get('role') not in roles:
                flash('Bạn không có quyền truy cập trang này.')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return wrapper
    return deco

# auth
@app.route('/register', methods=['GET','POST'])
def register():
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        address = request.form.get('address')
        password = request.form.get('password')

        session_db = Session()
        try:
            # kiểm tra email
            if session_db.query(User).filter(User.email == email).first():
                flash('Email đã được sử dụng.')
                return redirect(url_for('register'))

            # tạo user
            u = User(
                full_name=full_name,
                email=email,
                phone=phone,
                address=address,
                password_hash=generate_password_hash(password),
                role='customer'
            )

            # tạo customer
            c = Customer(
                full_name=full_name,
                email=email,
                phone=phone,
                address=address
            )

            session_db.add(u)
            session_db.add(c)
            session_db.commit()

            flash('Đăng ký thành công. Hãy đăng nhập.')
            return redirect(url_for('login'))

        finally:
            session_db.close()

    return render_template('register.html')


@app.route("/faq")
def faq():
    return render_template("support_faq.html")

@app.route("/warranty")
def warranty():
    return render_template("support_warranty.html")

@app.route("/terms")
def terms():
    return render_template("support_terms.html")

@app.route("/privacy")
def privacy():
    return render_template("support_privacy.html")


@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')

        session_db = Session()
        try:
            u = session_db.query(User).filter(User.email == email).first()

            if not u or not check_password_hash(u.password_hash, password):
                flash('Email hoặc mật khẩu không đúng.')
                return redirect(url_for('login'))


            session.clear()
            session['user_id'] = u.id
            session['full_name'] = u.full_name
            session['role'] = u.role



            if u.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif u.role == 'technician':
                return redirect(url_for('technician'))
            else:
                return redirect(url_for('index'))

        finally:
            session_db.close()

    return render_template('login.html')



@app.route('/submit_review', methods=['POST'])
@login_required
def submit_review():
    order_id = request.form.get('order_id')
    rating = request.form.get('rating')
    comment = request.form.get('comment', '')
    db = Session()
    try:

        order = db.query(WorkOrder).get(order_id)
        if not order:
            flash('Không tìm thấy đơn để đánh giá.', 'danger')
            return redirect(url_for('my_orders'))


        note = f"Đánh giá: {rating} sao. Nhận xét: {comment}"
        hist = OrderHistory(order_id=order.id, old_status=order.status, new_status=order.status, changed_by=session['user_id'], note=note)
        db.add(hist)
        db.commit()
        flash('Cảm ơn bạn đã đánh giá!', 'success')
    except Exception as e:
        db.rollback()
        flash('Lỗi khi gửi đánh giá: '+str(e), 'danger')
    finally:
        db.close()
    return redirect(url_for('my_orders'))


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


@app.route('/maintenance')
def maintenance():
    return render_template("customer_choice=.html")

@app.route('/tech_reports')
def tech_reports():
    db = SessionLocal()
    tech_id = session.get('user_id')  # technician đã đăng nhập
    if not tech_id:
        return "Bạn chưa đăng nhập", 403


    reports = db.query(WorkOrder).filter(
        WorkOrder.technician_id == tech_id,
        WorkOrder.status == 'da_hoan_thanh'
    ).order_by(WorkOrder.created_at.desc()).all()


    technician = db.query(User).filter(User.id == tech_id).first()

    return render_template('tech_reports.html', reports=reports, technician=technician)




@app.route('/')
def index():
    return render_template('index.html')


@app.route('/customer')
def customer_choice():
    return render_template('customer_choice.html')

@app.route('/customer/form/<order_type>', methods=['GET','POST'])
@login_required
def customer_form(order_type):
    if order_type not in ('baotri','suachua'):
        flash('Loại không hợp lệ')
        return redirect(url_for('customer_choice'))

    db = Session()
    try:

        uid = session.get("user_id")
        user = db.query(User).get(uid)

        if not user:
            flash("Không tìm thấy tài khoản người dùng!")
            return redirect(url_for('customer_choice'))


        customer = db.query(Customer).filter(Customer.email == user.email).first()

        if not customer:
            customer = Customer(
                full_name=user.full_name,
                phone=user.phone,
                email=user.email,
                address=user.address
            )
            db.add(customer)
            db.flush()

        if request.method == 'POST':
            description = request.form.get('description')
            machine_type = request.form.get('machine_type')


            customer.full_name = user.full_name
            customer.phone = user.phone
            customer.address = user.address
            db.add(customer)

            # --- TẠO ĐƠN ---
            order = WorkOrder(
                customer_id=customer.id,
                order_type=order_type,
                description=description,
                machine_type=machine_type
            )
            db.add(order)
            db.flush()


            files = request.files.getlist('images')

            for f in files:
                if not f or f.filename == '':
                    continue
                if not allowed_file(f.filename):
                    continue

                safe = secure_filename(f.filename)
                prefix = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
                final_name = f"{prefix}_{safe}"

                save_path = os.path.join(app.config['UPLOAD_FOLDER'], final_name)
                f.save(save_path)

                img = OrderImage(
                    order_id=order.id,
                    file_path=final_name,
                    uploaded_by=uid
                )
                db.add(img)

            db.commit()
            flash(f'Đã tạo đơn thành công! ')
            return redirect(url_for('customer_choice'))

        return render_template(
            'customer_form.html',
            order_type=order_type,
            customer=customer
        )

    except Exception as e:
        db.rollback()
        flash("Lỗi khi tạo đơn: " + str(e), 'danger')
        return redirect(url_for('customer_choice'))
    finally:
        db.close()




@app.route('/my/orders')
@login_required
def my_orders():
    uid = session['user_id']
    db = Session()
    try:
        user = db.query(User).get(uid)
        cust = None

        if user and user.email:
            cust = db.query(Customer).filter(Customer.email == user.email).first()

        orders = []
        if cust:
            orders = db.query(WorkOrder).filter(
                WorkOrder.customer_id == cust.id,
                WorkOrder.progress != -1
            ).order_by(WorkOrder.created_at.asc()).all()


            index = 1
            for o in orders:
                o.stt = index
                index += 1

        return render_template('my_orders.html', orders=orders)
    finally:
        db.close()



@app.route("/order/<int:order_id>/cancel", methods=["POST"])
@login_required
def customer_cancel_order(order_id):
    reason = request.form.get("reason", "").strip()
    db = Session()
    try:
        uid = session.get("user_id")
        user = db.query(User).get(uid)

        if not user:
            flash("Bạn chưa đăng nhập", "danger")
            return redirect(url_for("login"))


        customer = db.query(Customer).filter(
            Customer.email == user.email
        ).first()

        if not customer:
            flash("Không tìm thấy khách hàng", "danger")
            return redirect(url_for("my_orders"))


        order = db.query(WorkOrder).filter(
            WorkOrder.id == order_id,
            WorkOrder.customer_id == customer.id
        ).first()

        if not order:
            flash("Không tìm thấy đơn", "danger")
            return redirect(url_for("my_orders"))

        if order.status != "moi":
            flash("Chỉ được hủy đơn MỚI", "warning")
            return redirect(url_for("my_orders"))

        order.progress = -1
        order.cancel_reason = reason
        order.cancel_at = datetime.utcnow()

        db.commit()
        flash(f"Đã hủy đơn ", "success")

    except Exception as e:
        db.rollback()
        flash(f"Lỗi: {e}", "danger")
    finally:
        db.close()

    return redirect(url_for("my_orders"))


@app.route('/technician')
@role_required('technician', 'admin')
def technician():
    filter_status = request.args.get('status', 'all')

    db = Session()
    try:
        q = db.query(WorkOrder).options(
            joinedload(WorkOrder.customer),
            joinedload(WorkOrder.technician)
        )


        if session.get('role') == 'technician':
            q = q.filter(WorkOrder.technician_id == session['user_id'])


        if filter_status != 'all':
            q = q.filter(WorkOrder.status == filter_status)

        orders = q.order_by(

            case(
                (WorkOrder.status == 'dang_xu_ly', 1),
                (WorkOrder.status == 'da_hoan_thanh', 2),
                else_=3
            ),

            desc(WorkOrder.progress),

            WorkOrder.created_at.desc()
        ).all()


        all_orders_q = db.query(WorkOrder)

        if session.get('role') == 'technician':
            all_orders_q = all_orders_q.filter(
                WorkOrder.technician_id == session['user_id']
            )

        all_orders = all_orders_q.all()

        stats = {
            'total': len(all_orders),
            'new': len([o for o in all_orders if o.status == 'moi']),
            'processing': len([o for o in all_orders if o.status == 'dang_xu_ly']),
            'done': len([o for o in all_orders if o.status == 'da_hoan_thanh'])
        }

        technicians = db.query(User)\
                        .filter(User.role == "technician")\
                        .all()

        return render_template(
            "tech_orders.html",
            orders=orders,
            technicians=technicians,
            current_status=filter_status,
            stats=stats
        )
    finally:
        db.close()


@app.route('/technician/completed')
@role_required('technician', 'admin')
def technician_completed_orders():
    db = Session()
    try:
        q = db.query(WorkOrder).options(
            joinedload(WorkOrder.customer)
        ).filter(WorkOrder.status == 'da_hoan_thanh')

        if session.get('role') == 'technician':
            q = q.filter(WorkOrder.technician_id == session['user_id'])

        orders = q.order_by(WorkOrder.created_at.desc()).all()

        return render_template(
            'tech_completed_orders.html',
            orders=orders
        )
    finally:
        db.close()

@app.route("/technician/skills")
def technician_skills():
    db = SessionLocal()

    tech_id = session.get("user_id")
    if not tech_id:
        return "Bạn chưa đăng nhập", 403

    technician = db.query(User).filter(User.id == tech_id).first()

    skills = db.query(TechnicianSkill).filter(
        TechnicianSkill.technician_id == tech_id
    ).all()

    return render_template(
        "technician_skills.html",
        technician=technician,
        skills=skills
    )



@app.route('/order/<int:order_id>/update-price', methods=['POST'])
@role_required('technician', 'admin')
def update_order_price(order_id):
    price = request.form.get('price')

    if not price:
        flash("Vui lòng nhập giá tiền", "danger")
        return redirect(url_for('technician_completed_orders'))

    db = Session()
    try:
        order = db.query(WorkOrder).get(order_id)

        if not order:
            flash("Không tìm thấy đơn", "danger")
            return redirect(url_for('technician_completed_orders'))

        order.price = int(price)
        db.commit()

        flash("Đã lưu giá tiền thành công", "success")
        return redirect(url_for('technician_completed_orders'))
    finally:
        db.close()


@app.route('/admin')
@role_required('admin')
def admin_dashboard():
    db = Session()
    try:
        orders = db.query(WorkOrder).order_by(WorkOrder.created_at.desc()).all()

        technicians = []
        all_techs = db.query(User).filter(User.role == "technician").all()

        for t in all_techs:
            active = db.query(WorkOrder).filter(
                WorkOrder.technician_id == t.id,
                WorkOrder.status != "da_hoan_thanh"
            ).count()

            t.active_orders = active
            technicians.append(t)

        return render_template(
            "admin_orders.html",
            orders=orders,
            technicians=technicians
        )
    finally:
        db.close()


@app.route("/admin/accounts")
@role_required('admin')
def admin_accounts():
    db = Session()
    try:
        filter_role = request.args.get("role")

        if filter_role and filter_role.strip() != "":
            users = db.query(User).filter(User.role == filter_role).order_by(User.id.desc()).all()
        else:
            users = db.query(User).order_by(User.id.desc()).all()
            filter_role = ""

        return render_template(
            "admin_accounts.html",
            users=users,
            filter_role=filter_role,
            tab="accounts"
        )
    finally:
        db.close()




@app.route("/admin/update_role/<int:user_id>", methods=["POST"])
@role_required('admin')
def admin_update_role(user_id):
    new_role = request.form.get("role")

    db = Session()
    user = db.query(User).get(user_id)

    if not user:
        flash("Không tìm thấy tài khoản.")
        return redirect(url_for("admin_accounts"))

    user.role = new_role
    db.commit()
    db.close()

    flash("Cập nhật quyền thành công!")
    return redirect(url_for("admin_accounts"))


@app.route("/admin/orders")
@role_required("admin")
def admin_orders():
    order_type = request.args.get("type")
    status = request.args.get("status")
    group = request.args.get("group")

    db = Session()
    try:
        q = db.query(WorkOrder).filter(
            WorkOrder.progress != -1
        )

        # 🔹 Lọc loại đơn
        if order_type in ("baotri", "suachua"):
            q = q.filter(WorkOrder.order_type == order_type)

        # 🔹 Lọc trạng thái
        if status in ("moi", "dang_xu_ly", "da_hoan_thanh"):
            q = q.filter(WorkOrder.status == status)

        # 🔹 Lọc nhóm máy (machine_type)
        if group:
            q = q.filter(WorkOrder.machine_type == group)

        status_order = case(
            (WorkOrder.status == "moi", 1),
            (WorkOrder.status == "dang_xu_ly", 2),
            (WorkOrder.status == "da_hoan_thanh", 3),
            else_=4
        )

        # ===== PAGINATION =====
        page = request.args.get("page", 1, type=int)
        per_page = 20

        q = q.order_by(status_order, WorkOrder.created_at.desc())

        total = q.count()
        total_pages = (total + per_page - 1) // per_page

        orders = q.offset((page - 1) * per_page).limit(per_page).all()

        machine_groups = (
            db.query(WorkOrder.machine_type)
            .filter(WorkOrder.progress != -1)
            .distinct()
            .order_by(WorkOrder.machine_type)
            .all()
        )

        machine_groups = [m[0] for m in machine_groups if m[0]]

        # 🔹 Kỹ thuật viên
        technicians = []
        all_techs = db.query(User).filter(User.role == "technician").all()

        for t in all_techs:
            active = db.query(WorkOrder).filter(
                WorkOrder.technician_id == t.id,
                WorkOrder.status != "da_hoan_thanh",
                WorkOrder.progress != -1
            ).count()
            t.active_orders = active
            technicians.append(t)

        return render_template(
            "admin_orders.html",
            orders=orders,
            technicians=technicians,
            machine_groups=machine_groups,
            filter_type=order_type,
            filter_status=status,
            filter_group=group,
            page=page,
            total_pages=total_pages
        )


    finally:
        db.close()





# Order detail
@app.route('/order/<int:order_id>', methods=['GET', 'POST'])
def order_detail(order_id):
    db = Session()
    try:
        order = db.query(WorkOrder).get(order_id)
        if not order:
            flash("Không tìm thấy đơn hàng.", "danger")
            return redirect(url_for('technician_orders'))

        images = db.query(OrderImage).filter_by(order_id=order_id).all()

        history = db.query(OrderHistory)\
            .filter_by(order_id=order.id)\
            .order_by(OrderHistory.created_at.desc())\
            .all()

        technicians = db.query(User)\
            .filter(User.role == "technician")\
            .all()

        if request.method == 'POST':
            if request.form.get("action") == "update_status":
                old_status = order.status
                order.status = request.form.get("status")

                note = request.form.get("note")
                if note is not None:
                    note = note.strip()
                if note == "":
                    note = None

                hist = OrderHistory(
                    order_id=order.id,
                    old_status=old_status,
                    new_status=order.status,
                    changed_by=session.get("user_id"),
                    note=note
                )

                db.add(hist)
                db.commit()

                flash("Cập nhật trạng thái thành công!", "success")
                return redirect(url_for('order_detail', order_id=order.id))

        return render_template(
            "order_detail.html",
            order=order,
            images=images,
            history=history,
            technicians=technicians
        )

    finally:
        db.close()



def allowed_file(filename):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in ALLOWED_EXT


@app.route("/admin/assign/<int:order_id>", methods=["POST"])
@role_required("admin")
def admin_assign(order_id):
    tech_id = request.form.get("technician_id")

    db = Session()
    try:
        order = db.query(WorkOrder).get(order_id)

        if not order:
            flash("Không tìm thấy đơn.", "danger")
            return redirect(url_for("admin_orders"))


        if not tech_id:
            order.technician_id = None


            order.status = "moi"
            order.progress = 0

            db.commit()
            flash("Đã bỏ gán kỹ thuật viên, đơn trở về trạng thái MỚI.", "warning")
            return redirect(url_for("admin_orders"))


        tech_id = int(tech_id)

        active_orders = db.query(WorkOrder).filter(
            WorkOrder.technician_id == tech_id,
            WorkOrder.status != "da_hoan_thanh"
        ).count()

        MAX_ORDERS = 10

        if active_orders >= MAX_ORDERS:
            flash(
                f"Kỹ thuật viên đã có {active_orders}/{MAX_ORDERS} đơn!",
                "danger"
            )
            return redirect(url_for("admin_orders"))

        order.technician_id = tech_id


        order.status = "dang_xu_ly"

        db.commit()
        flash("Gán kỹ thuật viên thành công!", "success")

    except Exception as e:
        db.rollback()
        flash(f"Lỗi: {e}", "danger")
    finally:
        db.close()

    return redirect(url_for("admin_orders"))



@app.route("/admin/orders/cancel/<int:order_id>", methods=["POST"])
@role_required("admin")
def admin_cancel_order(order_id):
    db = Session()
    try:
        order = db.query(WorkOrder).get(order_id)
        if order:
            db.delete(order)
            db.commit()

        return redirect(url_for("admin_orders"))
    finally:
        db.close()


# Admin: change user role
@app.route('/admin/change_role/<int:user_id>', methods=['POST'])
@role_required('admin')
def admin_change_role(user_id):
    new_role = request.form.get('role')
    if new_role not in ('customer','technician','admin'):
        flash('Role không hợp lệ.')
        return redirect(url_for('admin_dashboard'))
    db = Session()
    try:
        u = db.query(User).get(user_id)
        if u:
            u.role = new_role
            db.commit()
            flash('Đã cập nhật quyền.')
    finally:
        db.close()
    return redirect(url_for('admin_dashboard'))


from collections import defaultdict
from datetime import datetime
import json

@app.route("/admin/reports")
@role_required("admin")
def admin_reports():
    db = Session()

    try:

        start = request.args.get("start")   # yyyy-mm-dd
        end = request.args.get("end")

        q = db.query(WorkOrder).options(
            joinedload(WorkOrder.customer),
            joinedload(WorkOrder.technician)
        )

        if start:
            q = q.filter(WorkOrder.created_at >= start)
        if end:
            q = q.filter(WorkOrder.created_at <= end)

        orders = q.all()


        total_orders = len(orders)
        total_processing = sum(1 for o in orders if o.status == "dang_xu_ly")
        total_done = sum(1 for o in orders if o.status == "da_hoan_thanh")
        total_revenue = sum(o.price or 0 for o in orders if o.status == "da_hoan_thanh")
        avg_revenue = int(total_revenue / total_done) if total_done else 0




        completion_rate = round((total_done / total_orders) * 100, 1) if total_orders else 0


        monthly_orders = defaultdict(int)
        monthly_revenue = defaultdict(int)

        for o in orders:
            if o.created_at:
                key = o.created_at.strftime("%m/%Y")
                monthly_orders[key] += 1
                if o.status == "da_hoan_thanh":
                    monthly_revenue[key] += o.price or 0

        labels = sorted(monthly_orders.keys(), key=lambda x: datetime.strptime(x, "%m/%Y"))

        monthly_data = json.dumps({
            "labels": labels,
            "datasets": [{
                "label": "Số đơn",
                "data": [monthly_orders[m] for m in labels]
            }]
        })

        revenue_data = json.dumps({
            "labels": labels,
            "datasets": [{
                "label": "Doanh thu (₫)",
                "data": [monthly_revenue[m] for m in labels]
            }]
        })


        growth = []
        for i in range(1, len(labels)):
            prev = monthly_revenue[labels[i-1]]
            cur = monthly_revenue[labels[i]]
            rate = round((cur - prev) / prev * 100, 1) if prev else 0
            growth.append(rate)


        bao_tri = sum(1 for o in orders if o.order_type == "bao_tri")
        sua_chua = sum(1 for o in orders if o.order_type == "sua_chua")

        type_data = json.dumps({
            "labels": ["Bảo trì", "Sửa chữa"],
            "datasets": [{"data": [bao_tri, sua_chua]}]
        })


        tech_stats = defaultdict(lambda: {
            "orders": 0,
            "done": 0,
            "revenue": 0
        })

        for o in orders:
            if o.technician:
                name = o.technician.full_name
                tech_stats[name]["orders"] += 1

                if o.status == "da_hoan_thanh":
                    tech_stats[name]["done"] += 1
                    tech_stats[name]["revenue"] += o.price or 0

        tech_labels = list(tech_stats.keys())
        tech_orders = [tech_stats[t]["orders"] for t in tech_labels]
        tech_revenue = [tech_stats[t]["revenue"] for t in tech_labels]

        tech_data = json.dumps({
            "labels": tech_labels,
            "datasets": [{
                "label": "Số đơn",
                "data": tech_orders
            }]
        })

        tech_revenue_data = json.dumps({
            "labels": tech_labels,
            "datasets": [{
                "label": "Doanh thu (₫)",
                "data": tech_revenue
            }]
        })


        tech_table = []
        for name, v in tech_stats.items():
            completion = round((v["done"] / v["orders"]) * 100, 1) if v["orders"] else 0
            avg = int(v["revenue"] / v["done"]) if v["done"] else 0

            tech_table.append({
                "name": name,
                "orders": v["orders"],
                "done": v["done"],
                "completion": completion,
                "revenue": v["revenue"],
                "avg": avg
            })

        tech_table.sort(key=lambda x: x["revenue"], reverse=True)


        done_orders = [o for o in orders if o.status == "da_hoan_thanh"]
        done_orders.sort(key=lambda x: x.created_at or datetime.min, reverse=True)

        return render_template(
            "admin_reports.html",
            tab="reports",

            # KPI
            total_orders=total_orders,
            total_processing=total_processing,
            total_done=total_done,
            total_revenue=total_revenue,
            avg_revenue=avg_revenue,
            completion_rate=completion_rate,

            # Charts
            monthly_data=monthly_data,
            revenue_data=revenue_data,
            type_data=type_data,
            tech_data=tech_data,
            tech_revenue_data=tech_revenue_data,

            # Pro data
            growth=growth,
            tech_table=tech_table,
            done_orders=done_orders,

            # filter
            start=start,
            end=end
        )
    finally:
        db.close()



@app.route('/admin/reports/export/excel')
@role_required('admin')
def export_reports_excel():
    db = Session()

    orders = (
        db.query(WorkOrder)
        .options(
            joinedload(WorkOrder.customer),
            joinedload(WorkOrder.technician)
        )
        .filter(
            WorkOrder.status == 'da_hoan_thanh',
            WorkOrder.price != None,
            WorkOrder.price > 0
        )
        .order_by(WorkOrder.created_at.desc())
        .all()
    )

    wb = Workbook()


    ws1 = wb.active
    ws1.title = "Chi tiết đơn"

    headers = ["ID", "Ngày", "Khách hàng", "Kỹ thuật viên", "Loại đơn", "Giá (VND)"]
    ws1.append(headers)

    for c in range(1, len(headers)+1):
        ws1.cell(row=1, column=c).font = Font(bold=True)

    total_revenue = 0

    for o in orders:
        price = int(o.price or 0)
        total_revenue += price

        ws1.append([
            o.id,
            o.created_at.strftime('%d/%m/%Y') if o.created_at else "",
            o.customer.full_name if o.customer else "",
            o.technician.full_name if o.technician else "",
            "Bảo trì" if o.order_type == "bao_tri" else "Sửa chữa",
            price
        ])

    ws1.append([])
    ws1.append(["", "", "", "", "TỔNG DOANH THU", total_revenue])


    ws2 = wb.create_sheet("Kỹ thuật viên")

    ws2.append(["Kỹ thuật viên", "Số đơn", "Đơn hoàn thành", "Doanh thu", "Tỷ lệ hoàn thành (%)"])
    for c in range(1, 6):
        ws2.cell(row=1, column=c).font = Font(bold=True)

    tech_stats = defaultdict(lambda: {"orders": 0, "done": 0, "revenue": 0})

    for o in orders:
        if o.technician:
            name = o.technician.full_name
            tech_stats[name]["orders"] += 1
            tech_stats[name]["done"] += 1
            tech_stats[name]["revenue"] += int(o.price or 0)

    for name, v in tech_stats.items():
        completion = round((v["done"] / v["orders"]) * 100, 1) if v["orders"] else 0
        ws2.append([name, v["orders"], v["done"], v["revenue"], completion])


    ws3 = wb.create_sheet("Tổng hợp")

    total_orders = len(orders)
    avg = int(total_revenue / total_orders) if total_orders else 0

    ws3.append(["Chỉ số", "Giá trị"])
    ws3.cell(1,1).font = Font(bold=True)
    ws3.cell(1,2).font = Font(bold=True)

    ws3.append(["Tổng đơn hoàn thành", total_orders])
    ws3.append(["Tổng doanh thu", total_revenue])
    ws3.append(["Doanh thu trung bình / đơn", avg])


    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="bao_cao_quan_tri.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/delivery')
@role_required('delivery')
def delivery_dashboard():
    db = Session()
    try:

        orders = db.query(WorkOrder).filter(
            WorkOrder.status.in_(["moi", "cho_giao_hang", "cho_nhan_lai"])
        ).order_by(WorkOrder.created_at.desc()).all()

        return render_template('delivery_orders.html', orders=orders)
    finally:
        db.close()

@app.route('/delivery/update/<int:order_id>', methods=['POST'])
@role_required('delivery')
def delivery_update(order_id):
    new_status = request.form.get('status')

    db = Session()
    try:
        order = db.query(WorkOrder).get(order_id)
        if not order:
            flash("Không tìm thấy đơn.", "danger")
            return redirect(url_for('delivery_dashboard'))

        old = order.status
        order.status = new_status

        hist = OrderHistory(
            order_id=order.id,
            old_status=old,
            new_status=new_status,
            changed_by=session['user_id'],
            note=f"Nhân viên giao cập nhật: {new_status}"
        )
        db.add(hist)
        db.commit()

        flash("Đã cập nhật trạng thái.")
    finally:
        db.close()

    return redirect(url_for('delivery_dashboard'))


@app.route('/admin/update/<int:order_id>', methods=['POST'])
@role_required('admin')
def admin_update_status(order_id):
    status = request.form.get('status')
    db = Session()
    try:
        order = db.query(WorkOrder).get(order_id)
        if order:
            order.status = status
            db.commit()
            flash('Admin đã cập nhật trạng thái.')
    finally:
        db.close()
    return redirect(url_for('admin_dashboard'))

def technician_available(db, tech_id, limit=10):
    active_orders = db.query(WorkOrder).filter(
        WorkOrder.technician_id == tech_id,
        WorkOrder.status != "da_hoan_thanh"
    ).count()

    return active_orders < limit

@app.route('/order/<int:order_id>/progress', methods=['POST'])
def update_order_progress(order_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    progress = int(request.form.get('progress', 0))

    db = Session()
    order = db.query(WorkOrder).get(order_id)

    order.progress = progress

    if progress == 0:
        order.status = 'moi'
    elif progress < 100:
        order.status = 'dang_xu_ly'
    else:
        order.progress = 100
        order.status = 'da_hoan_thanh'

    db.commit()
    db.close()


    return redirect(request.referrer)


if __name__ == '__main__':
    app.run(debug=True)


